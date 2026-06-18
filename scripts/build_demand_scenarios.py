#!/usr/bin/env python
"""
Build the four CORSIA demand scenarios from the v2.3 demand workbook.

Reads  Demand/SAF_CORSIA_Demand_Model_v2.3.xlsx  and writes, for each scenario
(C1 A&M base, C2 A&M policy-driven, C3 A&M climate push, C4 IATA reference):

  * a route-targets flight_routes.csv  (per-scenario ROUTES sheet → model schema)
  * a wtp_params.csv                   (per-scenario WTP ceilings + carbon price)
  * a shared aircraft_types.csv        (union of all aircraft types in the routes)
  * the remaining input CSVs           (copied from the current data/mock/)
  * meta.json                          (run settings: route_targets, efficiency rate, …)

into  scenarios/<name>/, and refreshes the working data/mock/ + data/templates/
defaults to the C1 base case.

Re-runnable: regenerate after editing the workbook. One-off data tooling — not
imported by the app.
"""

from __future__ import annotations

import csv
import json
import os
import shutil
from datetime import datetime, timezone

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
_XLSX = os.path.join(_ROOT, "Demand", "SAF_CORSIA_Demand_Model_v2.3.xlsx")
_MOCK = os.path.join(_ROOT, "data", "mock")
_TMPL = os.path.join(_ROOT, "data", "templates")
_SCEN = os.path.join(_ROOT, "scenarios")

REGIONS = ["EU", "US", "APAC", "MENA", "LATAM", "ROW"]
YEARS = list(range(2025, 2051))

# Model flight_routes.csv schema (column order matches the existing mock file).
ROUTE_COLS = [
    "route_id", "airline_id", "operator_name", "segment",
    "origin_airport", "origin_country", "origin_region",
    "dest_airport", "dest_country", "dest_region",
    "flight_type", "aircraft_type", "annual_flights_2025", "distance_km",
    "annual_growth_rate",
    "saf_pct_2025", "saf_pct_2030", "saf_pct_2035",
    "saf_pct_2040", "saf_pct_2045", "saf_pct_2050",
]
# Excel column indices (0-based) feeding each model column.
_XL = dict(route_id=0, airline_id=1, operator_name=2, segment=3,
           origin_airport=4, origin_country=5, origin_region=6,
           dest_airport=7, dest_country=8, dest_region=9,
           flight_type=10, aircraft_type=11, annual_flights_2025=12,
           distance_km=14, annual_growth_rate=16,
           saf_pct_2025=24, saf_pct_2030=26, saf_pct_2035=28,
           saf_pct_2040=30, saf_pct_2045=32, saf_pct_2050=34)
_EFF_COL = 17  # "Fuel Eff. (t/km)" — authoritative per-route efficiency

# scenario key → (folder name, route sheet, case3 endpoints, carbon 2050, eff rate)
SCENARIOS = {
    "C1": ("C1_AM_base", "ROUTES_C1",
           {"EU": (2600, 3200), "US": (1500, 2200), "APAC": (1300, 1800), "MENA": (1000, 1300)},
           190, 0.0),
    "C2": ("C2_AM_policy", "ROUTES_C2",
           {"EU": (2800, 3600), "US": (1700, 2400), "APAC": (1500, 2000), "MENA": (1100, 1400)},
           250, 0.015),
    "C3": ("C3_AM_climate", "ROUTES_C3",
           {"EU": (3000, 4000), "US": (1700, 2500), "APAC": (1600, 2200), "MENA": (1200, 1600)},
           300, 0.02),
    "C4": ("C4_IATA_reference", "ROUTES_IATA",
           {"EU": (3000, 4000), "US": (1700, 2500), "APAC": (1600, 2200), "MENA": (1200, 1600)},
           300, 0.0),
}


def _f(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _interp(y, v0, v1):
    """Linear value at year y between 2025 (v0) and 2050 (v1)."""
    return v0 + (v1 - v0) * (y - 2025) / (2050 - 2025)


def _route_rows(ws):
    """Yield valid route rows (Route ID starting with 'R')."""
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[_XL["route_id"]] and str(r[_XL["route_id"]]).strip().startswith("R"):
            yield r


def build_aircraft(wb) -> list[dict]:
    """Union of aircraft types across route sheets; efficiency from the route
    'Fuel Eff.' column, category from the AIRCRAFT sheet where available."""
    cat = {}
    for r in wb["AIRCRAFT"].iter_rows(min_row=4, values_only=True):
        if r[0] and r[1] is not None:
            cat[str(r[0]).strip()] = (str(r[2]).strip() if r[2] else "Unknown")
    eff: dict[str, float] = {}
    for _, sheet, *_ in SCENARIOS.values():
        for row in _route_rows(wb[sheet]):
            t = str(row[_XL["aircraft_type"]]).strip()
            e = _f(row[_EFF_COL], None)
            if t and e:
                eff.setdefault(t, e)
    return [{"aircraft_type": t,
             "fuel_efficiency_t_per_km": eff[t],
             "category": cat.get(t, "Unknown")}
            for t in sorted(eff)]


def convert_routes(ws) -> list[dict]:
    out = []
    for r in _route_rows(ws):
        row = {}
        for col in ROUTE_COLS:
            v = r[_XL[col]]
            if col == "annual_flights_2025":
                row[col] = int(round(_f(v)))
            elif col in ("origin_region", "dest_region", "origin_country",
                         "dest_country", "origin_airport", "dest_airport",
                         "flight_type", "aircraft_type", "route_id",
                         "airline_id", "operator_name", "segment"):
                row[col] = ("" if v is None else str(v).strip())
            else:
                row[col] = _f(v)
        out.append(row)
    return out


def _read_wtp_baseline() -> dict[tuple, dict]:
    """Current wtp_params.csv keyed by (year, region) — preserves jet fuel,
    IRR, mode, and the LATAM/ROW case3 trajectory the user asked to keep."""
    base = {}
    with open(os.path.join(_MOCK, "wtp_params.csv"), newline="") as fh:
        for d in csv.DictReader(fh):
            base[(int(d["year"]), d["region"])] = d
    return base


def build_wtp(baseline, case3_ends, carbon_2050) -> list[dict]:
    rows = []
    for y in YEARS:
        carbon = round(_interp(y, 30.0, carbon_2050), 2)
        for region in REGIONS:
            b = baseline[(y, region)]
            row = {
                "year": y, "region": region,
                "jet_fuel_price_usd_per_mt": b["jet_fuel_price_usd_per_mt"],
                "corsia_credit_usd_per_tco2": carbon,
                "case3_penalty_usd_per_mt": b["case3_penalty_usd_per_mt"],
                "target_irr_pct": b["target_irr_pct"],
                "wtp_mode": b["wtp_mode"],
            }
            if region in case3_ends:  # EU/US/APAC/MENA per table; LATAM/ROW kept as-is
                v0, v1 = case3_ends[region]
                row["case3_penalty_usd_per_mt"] = round(_interp(y, v0, v1), 2)
            rows.append(row)
    return rows


def _write_csv(path, fieldnames, rows):
    with open(path, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def main():
    wb = openpyxl.load_workbook(_XLSX, read_only=True, data_only=True)
    os.makedirs(_SCEN, exist_ok=True)

    aircraft = build_aircraft(wb)
    aircraft_fields = ["aircraft_type", "fuel_efficiency_t_per_km", "category"]
    baseline = _read_wtp_baseline()
    wtp_fields = ["year", "region", "jet_fuel_price_usd_per_mt",
                  "corsia_credit_usd_per_tco2", "case3_penalty_usd_per_mt",
                  "target_irr_pct", "wtp_mode"]

    # 13 snapshot CSVs (scenario_builder.INPUT_CSVS); these 10 are copied as-is.
    copy_as_is = [
        "airlines.csv", "corsia_schedule.csv", "corsia_suppression.csv",
        "national_blending_mandates.csv", "committed_capacity.csv",
        "refinery_capacity.csv", "domestic_supply_priority.csv",
        "feedstock_availability.csv", "transport_costs.csv", "regulatory_params.csv",
    ]
    now = datetime.now(timezone.utc).isoformat()
    c1_routes = c1_wtp = None

    for key, (folder, sheet, case3_ends, carbon_2050, eff_rate) in SCENARIOS.items():
        dest = os.path.join(_SCEN, folder)
        os.makedirs(dest, exist_ok=True)

        routes = convert_routes(wb[sheet])
        wtp = build_wtp(baseline, case3_ends, carbon_2050)

        _write_csv(os.path.join(dest, "flight_routes.csv"), ROUTE_COLS, routes)
        _write_csv(os.path.join(dest, "wtp_params.csv"), wtp_fields, wtp)
        _write_csv(os.path.join(dest, "aircraft_types.csv"), aircraft_fields, aircraft)
        for f in copy_as_is:
            shutil.copy(os.path.join(_MOCK, f), dest)

        meta = {
            "scenario_name": folder,
            "demand_mode": "route_targets",
            "include_domestic": False,
            "route_sample_fraction": 1.0,
            "demand_scale_factor": 1.0,
            "efficiency_improvement_rate": eff_rate,
            "start_year": 2025, "end_year": 2050,
            "run_completed": False,
            "saved_at": now,
            "source": "SAF_CORSIA_Demand_Model_v2.3.xlsx",
        }
        with open(os.path.join(dest, "meta.json"), "w") as fh:
            json.dump(meta, fh, indent=2)

        print(f"{folder:20} routes={len(routes)}  wtp_rows={len(wtp)}  eff={eff_rate}")
        if key == "C1":
            c1_routes, c1_wtp = routes, wtp

    # Refresh working defaults + templates to C1 (calibrated UN-bunker baseline).
    for target_dir in (_MOCK, _TMPL):
        _write_csv(os.path.join(target_dir, "flight_routes.csv"), ROUTE_COLS, c1_routes)
        _write_csv(os.path.join(target_dir, "wtp_params.csv"), wtp_fields, c1_wtp)
        _write_csv(os.path.join(target_dir, "aircraft_types.csv"), aircraft_fields, aircraft)

    print(f"\naircraft types: {len(aircraft)}  →  data/mock + data/templates set to C1")
    print(f"scenarios written under: {_SCEN}")


if __name__ == "__main__":
    main()
